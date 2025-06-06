from datetime import datetime

import numpy as np
import pandas as pd

import matplotlib.dates as mpd
import matplotlib.pylab as mpl
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.indexes import example_color

mpl.rcParams['font.sans-serif'] = ['FangSong']
mpl.rcParams['axes.unicode_minus'] = False
plt.switch_backend('TkAgg')
pd.set_option('display.max_rows', None)  # 显示所有行
pd.set_option('display.max_columns', None)  # 显示所有列


def calc(d1, d2):
    return round((d1 - d2) / d2 * 100, 2)


def apply_percent_styles(ws):
    for row in ws.iter_rows(min_row=2, min_col=2):  # 跳过标题行，并从第二列开始
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= 2:
                    cell.fill = PatternFill(start_color="ff7875", end_color="ff7875", fill_type="solid")  # 红色
                elif cell.value <= -2:
                    cell.fill = PatternFill(start_color="95de64", end_color="95de64", fill_type="solid")  # 绿色
                elif 2 > cell.value > 0:
                    cell.fill = PatternFill(start_color="ffccc7", end_color="ffccc7", fill_type="solid")  # 浅红色
                elif -2 < cell.value < 0:
                    cell.fill = PatternFill(start_color="d9f7be", end_color="d9f7be", fill_type="solid")  # 浅绿色

def apply_range_styles(ws):
    for row in ws.iter_rows(min_row=2, min_col=2):  # 跳过标题行，并从第二列开始
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= 20:
                    cell.fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")  # 红色
                elif 20 > cell.value >= 10:
                    cell.fill = PatternFill(start_color="ffa6a6", end_color="ffa6a6", fill_type="solid")  # 浅红色
                elif 10 > cell.value >= 0:
                    cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")  # 无


# 要对比的标的列表
symbol_list = [
    'SH510300',  # 沪深300
    'SH510500',  # 中证500
    # 'SH512100',  # 中证1000
    # 'SH516160',  # 新能源ETF
    # 'SH512890',  # 红利ETF
    # 'SH588000',  # 科创50
    # 'SZ159915',  # 创业板ETF
]

year_range = 12  # 时间
# date_type = 'W'  # 周度
date_type = 'ME'  # 月度
# date_type = 'QE'  # 季度
# date_type = 'YE'  # 年度

# 涨跌幅
def calc_month_percent():
    # end_date = pd.Timestamp("2021-12-31")
    end_date = pd.Timestamp.now()
    start_date = end_date - pd.DateOffset(years=year_range)
    aggs = []

    for symbol in symbol_list:
        df = pd.read_excel(f'../data/download_{symbol}.xlsx')
        df['date'] = pd.to_datetime(df['日期Date'], format='%Y%m%d')
        df = df[(df['date'] >= start_date) & (df['date'] <= end_date)]
        df.set_index('date', inplace=True)
        # 按月重采样，获取每月第一个交易日的开盘价和最后一个交易日的收盘价
        agg = df.resample(date_type).agg({
            #     '开盘Open': 'first',
            '收盘Close': 'last'
        })
        # # 计算涨跌幅
        agg[f'{symbol}涨幅'] = calc(agg['收盘Close'], agg['收盘Close'].shift(1))
        aggs.append(agg[[f'{symbol}涨幅']])

    merged_df = pd.concat(aggs, axis=1, join='inner')
    merged_df['Month'] = [d.strftime('%m') for d in merged_df.index]
    merged_df['Year'] = [d.strftime('%Y') for d in merged_df.index]

    # 打印结果查看
    # print("月涨跌幅对比:")

    output_path = f'../output/compare_etf_percent_{date_type}.xlsx'
    with pd.ExcelWriter(output_path) as writer:
        for col in merged_df.columns:
            if '涨幅' in col:
                cross_tab = pd.pivot_table(merged_df, values=col,
                                           index='Year', columns='Month', aggfunc='sum', fill_value=0)
                cross_tab.to_excel(writer, sheet_name=col)

    # 加载工作簿并应用样式
    wb = load_workbook(output_path)

    for col in merged_df.columns:
        if '涨幅' in col:
            apply_percent_styles(wb[col])

    # 保存修改后的工作簿
    wb.save(output_path)


# 范围差
def calc_month_range():
    # end_date = pd.Timestamp("2021-12-31")
    end_date = pd.Timestamp.now()
    start_date = end_date - pd.DateOffset(years=year_range)
    aggs = []

    for symbol in symbol_list:
        df = pd.read_excel(f'../data/download_{symbol}.xlsx')
        df['date'] = pd.to_datetime(df['日期Date'], format='%Y%m%d')
        df = df[(df['date'] >= start_date) & (df['date'] <= end_date)]
        df.set_index('date', inplace=True)
        # 按月重采样
        agg = df.resample(date_type).agg({
            '最高High': 'max',
            '最低Low': 'min'
        })
        # # 计算涨跌幅
        agg[f'{symbol}高低差'] = calc(agg['最高High'], agg['最低Low'])
        aggs.append(agg[[f'{symbol}高低差']])

    merged_df = pd.concat(aggs, axis=1, join='inner')
    merged_df['Month'] = [d.strftime('%m') for d in merged_df.index]
    merged_df['Year'] = [d.strftime('%Y') for d in merged_df.index]

    # 打印结果查看
    # print("月涨跌幅对比:")

    output_path = f'../output/compare_etf_range{date_type}.xlsx'
    with pd.ExcelWriter(output_path) as writer:
        for col in merged_df.columns:
            if '高低差' in col:
                cross_tab = pd.pivot_table(merged_df, values=col,
                                           index='Year', columns='Month', aggfunc='sum', fill_value=0)
                cross_tab.to_excel(writer, sheet_name=col)

    # 加载工作簿并应用样式
    wb = load_workbook(output_path)
    for col in merged_df.columns:
        if '高低差' in col:
            apply_range_styles(wb[col])

    # 保存修改后的工作簿
    wb.save(output_path)


# =SUM(B2:B14)/12
calc_month_percent()

# =SUM(B2:B14) / 12
# calc_month_range()
